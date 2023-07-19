from django.contrib.auth.models import User
from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render,redirect,get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook
from rest_framework import viewsets, permissions

from deportistas.forms import DeportistaFormulario
from deportistas.models import Deportista, Competencia
from deportistas.serializers import UserSerializer


#from deportistas.serializers import UserSerializer

# class UserViewSet(viewsets.ModelViewSet):
#     queryset = Competencia.objects.all().order_by('-date_joined')
#     serializer_class = UserSerializer
#     permission_classes = [permissions.IsAuthenticated]


# Create your views here.
def agregar_deportista(request):
    pagina = loader.get_template('agregar_deportista.html')
    if request.method == 'GET':
        formulario = DeportistaFormulario
    elif request.method =='POST':
        formulario = DeportistaFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos,request))


def ver_deportista (request, idDeportista):
    pagina = loader.get_template('ver_deportista.html')
    deportista = get_object_or_404(Deportista, pk=idDeportista)
    mensaje = {'deportista': deportista}
    return HttpResponse(pagina.render(mensaje, request))


def editar_deportista(request,idDeportista):
    pagina = loader.get_template('editar_deportista.html')
    deportista = get_object_or_404(Deportista, pk=idDeportista)
    if request.method == 'GET':
        formulario = DeportistaFormulario(instance=deportista)
    elif request.method =='POST':
        formulario = DeportistaFormulario(request.POST, instance=deportista)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    mensaje = {'formulario': formulario}
    return HttpResponse(pagina.render(mensaje, request))


def eliminar_deportista(request,idDeportista):
    deportista = get_object_or_404(Deportista, pk=idDeportista)
    if deportista:
        deportista.delete()
        return redirect('inicio')


def generar_reporte(request):
    # Obtenemos todas las personas de nuestra base de datos
    #deportistas = Deportista.objects.all()
    deportistas = Deportista.objects.order_by('apellido')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE DEPORTISTAS'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'CEDULA'
    ws['E3'] = 'SEXO'
    ws['F3'] = 'EDAD'
    ws['G3'] = 'DISCIPLINA'
    ws['H3'] = 'CIUDAD'
    ws['I3'] = 'SEDE'
    ws['J3'] = 'COMPETENCIA'
    ws['k3'] = 'REPRESENTANTE'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for deportista in deportistas:
        ws.cell(row=cont, column=2).value = deportista.nombre
        ws.cell(row=cont, column=3).value = deportista.apellido
        ws.cell(row=cont, column=4).value = deportista.cedula
        ws.cell(row=cont, column=5).value = deportista.sexo
        ws.cell(row=cont, column=6).value = deportista.edad
        ws.cell(row=cont, column=7).value = deportista.disciplina
        ws.cell(row=cont, column=8).value = deportista.ciudad
        ws.cell(row=cont, column=9).value = deportista.sede.nombre
        ws.cell(row=cont, column=10).value = deportista.competencia.categoria
        ws.cell(row=cont, column=11).value = deportista.representante.nombre_representante
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteDeportistasExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


class UserViewSet(viewsets.ModelViewSet):
    queryset = Competencia.objects.all().order_by('torneo')
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

